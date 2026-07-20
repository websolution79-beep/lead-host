from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path.home() / "AppData" / "Local" / "Temp" / "Elenco-comuni-italiani.xlsx"
OUT = ROOT / "src" / "lib" / "geo" / "italy-geo.ts"


def main() -> None:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    regions: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))

    for row in sheet.iter_rows(min_row=2, values_only=True):
        city = clean(row[6])
        region = clean(row[10])
        province = clean(row[11])

        if not city or not region or not province:
            continue

        regions[region][province].add(city)

    data = [
        {
            "region": region,
            "provinces": [
                {"province": province, "cities": sorted(cities, key=sort_key)}
                for province, cities in sorted(provinces.items(), key=lambda item: sort_key(item[0]))
            ],
        }
        for region, provinces in sorted(regions.items(), key=lambda item: sort_key(item[0]))
    ]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(
        "/* Generated from ISTAT 'Elenco dei comuni italiani' permalink, updated 2026-02-21. */\n"
        "export type ItalyGeoProvince = {\n"
        "  province: string;\n"
        "  cities: string[];\n"
        "};\n\n"
        "export type ItalyGeoRegion = {\n"
        "  region: string;\n"
        "  provinces: ItalyGeoProvince[];\n"
        "};\n\n"
        f"export const ITALY_GEO = {json.dumps(data, ensure_ascii=False, indent=2)} as const satisfies readonly ItalyGeoRegion[];\n",
        encoding="utf-8",
    )

    city_count = sum(len(province["cities"]) for region in data for province in region["provinces"])
    province_count = sum(len(region["provinces"]) for region in data)
    print(f"Generated {OUT}")
    print(f"Regions: {len(data)}")
    print(f"Provinces/statistical units: {province_count}")
    print(f"Cities: {city_count}")


def clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def sort_key(value: str) -> str:
    return value.casefold()


if __name__ == "__main__":
    main()
